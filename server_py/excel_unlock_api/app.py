from __future__ import annotations

import io
import os
from copy import copy

from flask import Flask, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Allow larger uploads (adjust as needed)
app.config["MAX_CONTENT_LENGTH"] = 150 * 1024 * 1024  # 150MB


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
    return response


@app.get("/health")
def health():
    return {"ok": True}


@app.post("/unlock")
def unlock_excel():
    if request.method == "OPTIONS":
        return "", 204

    if "file" not in request.files:
        return "Chyba: Žádný soubor nebyl nahrán", 400

    uploaded = request.files["file"]
    if not uploaded or uploaded.filename is None or uploaded.filename == "":
        return "Chyba: Prázdné jméno souboru", 400

    filename = secure_filename(uploaded.filename)
    _, ext = os.path.splitext(filename.lower())
    if ext not in (".xlsx", ".xlsm"):
        return "Chyba: Podporované jsou pouze soubory .xlsx a .xlsm", 400

    try:
        data = uploaded.read()
        if not data:
            return "Chyba: Soubor je prázdný", 400

        # keep_vba preserves macros for .xlsm; for .xlsx it is harmless but unnecessary.
        keep_vba = ext == ".xlsm"
        wb = load_workbook(io.BytesIO(data), keep_vba=keep_vba)

        for sheet in wb.worksheets:
            # Disable sheet protection
            sheet.protection.enabled = False

            # Optionally unlock cell styles. This is the behavior you validated in Python/openpyxl.
            # It can take longer on huge sheets, but preserves file size and formatting reliably.
            for row in sheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False, hidden=False)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        base = os.path.splitext(filename)[0]
        download_name = f"{base}-odemceno.xlsm"

        return send_file(
            out,
            mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
            as_attachment=True,
            download_name=download_name,
            max_age=0,
        )

    except Exception as e:
        return f"Chyba při zpracování: {str(e)}", 500


@app.post("/merge")
def merge_excel():
    if request.method == "OPTIONS":
        return "", 204

    if "file" not in request.files:
        return "Chyba: Žádný soubor nebyl nahrán", 400

    uploaded = request.files["file"]
    if not uploaded or uploaded.filename is None or uploaded.filename == "":
        return "Chyba: Prázdné jméno souboru", 400

    filename = secure_filename(uploaded.filename)
    _, ext = os.path.splitext(filename.lower())
    if ext not in (".xlsx", ".xlsm"):
        return "Chyba: Podporované jsou pouze soubory .xlsx a .xlsm", 400

    headers = [
        "List",
        "Výběrové řízení",
        "PČ",
        "Typ",
        "Kód",
        "Popis",
        "MJ",
        "Množství",
        "J.cena [CZK]",
        "Cena celkem [CZK]",
        "Cenová soustava",
    ]
    max_source_cols = len(headers) - 1  # source A..J -> target B..K (A is "List")
    skip_sheets = {"Rekapitulace stavby", "Pokyny pro vyplnění"}

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    sep_font = Font(bold=True, size=12, color="FFFFFF")
    sep_alignment = Alignment(horizontal="center", vertical="center")

    def copy_format(src, tgt):
        if not src.has_style:
            return
        tgt.font = copy(src.font)
        tgt.fill = copy(src.fill)
        tgt.border = copy(src.border)
        tgt.alignment = copy(src.alignment)
        tgt.number_format = src.number_format
        tgt.protection = copy(src.protection)

    try:
        data = uploaded.read()
        if not data:
            return "Chyba: Soubor je prázdný", 400

        # We always output .xlsx; macros are not preserved.
        source_wb = load_workbook(io.BytesIO(data))

        target_wb = Workbook()
        target_wb.remove(target_wb.active)
        combined = target_wb.create_sheet("Kombinovane")

        # Header row (Row 1)
        for idx, label in enumerate(headers, start=1):
            cell = combined.cell(row=1, column=idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        combined.freeze_panes = "A2"
        combined.row_dimensions[1].height = 18
        combined.column_dimensions["A"].width = 25

        col_widths = {}
        row = 2

        for sheet in source_wb.worksheets:
            if sheet.title in skip_sheets:
                continue
            if sheet.sheet_state in ("hidden", "veryHidden"):
                continue

            max_r = sheet.max_row or 0
            max_c = min(sheet.max_column or 0, max_source_cols)

            if max_r == 1 and max_c == 1 and sheet["A1"].value is None:
                continue

            # Separator row per sheet
            sep_cell = combined.cell(row=row, column=1, value=f"=== {sheet.title} ===")
            sep_cell.font = sep_font
            sep_cell.fill = header_fill
            sep_cell.alignment = sep_alignment
            for col in range(2, len(headers) + 1):
                c = combined.cell(row=row, column=col, value=None)
                c.font = sep_font
                c.fill = header_fill
                c.alignment = sep_alignment
            combined.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            combined.row_dimensions[row].height = 20
            row += 1

            data_start_row = row

            # Data copy
            for r in range(1, max_r + 1):
                target_r = data_start_row + r - 1
                combined.cell(row=target_r, column=1, value=sheet.title)
                for c in range(1, max_c + 1):
                    src_cell = sheet.cell(row=r, column=c)
                    if src_cell.value is None and not src_cell.has_style:
                        continue
                    tgt_cell = combined.cell(row=target_r, column=c + 1)
                    tgt_cell.value = src_cell.value
                    copy_format(src_cell, tgt_cell)

                if sheet.row_dimensions[r].height is not None:
                    combined.row_dimensions[target_r].height = sheet.row_dimensions[r].height

            # Merged cells (shifted by +1 column, and +data_start_row offset)
            for merged in sheet.merged_cells.ranges:
                min_row, min_col, max_row, max_col = (
                    merged.min_row,
                    merged.min_col,
                    merged.max_row,
                    merged.max_col,
                )
                if min_col > max_source_cols:
                    continue
                clipped_max_col = min(max_col, max_source_cols)
                if clipped_max_col < min_col:
                    continue
                combined.merge_cells(
                    start_row=data_start_row + min_row - 1,
                    start_column=min_col + 1,
                    end_row=data_start_row + max_row - 1,
                    end_column=clipped_max_col + 1,
                )

            # Column widths (keep max across sheets)
            for c in range(1, max_c + 1):
                src_col = get_column_letter(c)
                tgt_col = get_column_letter(c + 1)
                w = sheet.column_dimensions[src_col].width
                if w is None:
                    continue
                col_widths[tgt_col] = max(col_widths.get(tgt_col, 0), float(w))

            # Gap row between sheets
            row = data_start_row + max_r + 1

        for col, w in col_widths.items():
            combined.column_dimensions[col].width = w

        max_row_combined = combined.max_row
        combined.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row_combined}"

        out = io.BytesIO()
        target_wb.save(out)
        out.seek(0)

        base = os.path.splitext(filename)[0]
        download_name = f"{base}_combined_final.xlsx"
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
            max_age=0,
        )

    except Exception as e:
        return f"Chyba při zpracování: {str(e)}", 500


if __name__ == "__main__":
    port = int(os.environ.get("EXCEL_UNLOCK_PORT", "5000"))
    app.run(host="0.0.0.0", port=port)
