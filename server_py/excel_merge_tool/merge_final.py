#!/usr/bin/env python3
from __future__ import annotations

import os
import sys
from copy import copy
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SKIP_SHEETS = ("Rekapitulace stavby", "Pokyny pro vyplnění")

HEADERS = [
    "List",
    "Výběrové řízení",
    "PČ",
    "Typ",
    "Kód",
    "Popis",
    "MJ",
    "Množství",
    "J.cena [CZK]",
    "Cena celkem [CZK]",
    "Cenová soustava",
]

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

SHEET_HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
SHEET_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _should_skip_sheet(sheet) -> bool:
    title = (sheet.title or "").strip()
    if not title:
        return True
    if any(s.lower() == title.lower() for s in SKIP_SHEETS):
        return True
    if sheet.sheet_state in ("hidden", "veryHidden"):
        return True
    return False


def _iter_sheets_to_process(wb) -> Iterable:
    for name in wb.sheetnames:
        sheet = wb[name]
        if _should_skip_sheet(sheet):
            continue
        yield sheet


def copy_format(src, tgt) -> None:
    if not src.has_style:
        return
    tgt._style = copy(src._style)
    tgt.number_format = src.number_format
    tgt.protection = copy(src.protection)
    tgt.alignment = copy(src.alignment)


def merge_final(input_file: str, output_file: str) -> None:
    print(f"Loading: {input_file}")
    source_wb = openpyxl.load_workbook(input_file)
    sheets = list(_iter_sheets_to_process(source_wb))
    print(f"Processing {len(sheets)} sheets")

    target_wb = openpyxl.Workbook()
    target_wb.remove(target_wb.active)
    combined = target_wb.create_sheet("Kombinovane")

    # Header row (Row 1)
    for idx, label in enumerate(HEADERS, start=1):
        cell = combined.cell(row=1, column=idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    combined.freeze_panes = "A2"
    combined.column_dimensions["A"].width = 25

    row = 2
    max_source_cols = len(HEADERS) - 1  # we insert "List" in col A
    col_widths: dict[str, float] = {}

    for sheet in sheets:
        sheet_name = sheet.title
        print(f"Processing: {sheet_name}")

        max_r = sheet.max_row or 0
        max_c = min(sheet.max_column or 0, max_source_cols)

        if max_r <= 1 and max_c <= 1 and sheet["A1"].value is None:
            continue

        # Blue separator row per sheet
        header_cell = combined.cell(row=row, column=1, value=f"=== {sheet_name} ===")
        header_cell.font = SHEET_HEADER_FONT
        header_cell.fill = HEADER_FILL
        header_cell.alignment = SHEET_HEADER_ALIGNMENT
        combined.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        row += 1

        # Copy data (row by row) while preserving formatting
        # Column A is the source sheet name. Source columns map to B..K (10 columns).
        merge_ranges = list(sheet.merged_cells.ranges)
        row_offset = row - 1  # source row 1 -> target row (row_offset + 1)

        for r in range(1, max_r + 1):
            target_r = row_offset + r
            combined.cell(row=target_r, column=1, value=sheet_name)

            for c in range(1, max_c + 1):
                src_cell = sheet.cell(row=r, column=c)
                if src_cell.value is None and not src_cell.has_style:
                    continue
                tgt_cell = combined.cell(row=target_r, column=c + 1)
                tgt_cell.value = src_cell.value
                copy_format(src_cell, tgt_cell)

            # Preserve row height
            if sheet.row_dimensions[r].height is not None:
                combined.row_dimensions[target_r].height = sheet.row_dimensions[r].height

        # Preserve merged cells (shifted by +1 column, +row_offset rows)
        for merged in merge_ranges:
            min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
            # Only merge within copied range (clip to max_c)
            if min_col > max_source_cols:
                continue
            clipped_max_col = min(max_col, max_source_cols)
            combined.merge_cells(
                start_row=row_offset + min_row,
                start_column=min_col + 1,
                end_row=row_offset + max_row,
                end_column=clipped_max_col + 1,
            )

        # Column widths: use max width across all processed sheets
        for c in range(1, max_c + 1):
            src_col_letter = get_column_letter(c)
            tgt_col_letter = get_column_letter(c + 1)
            w = sheet.column_dimensions[src_col_letter].width
            if w is None:
                continue
            col_widths[tgt_col_letter] = max(col_widths.get(tgt_col_letter, 0), float(w))

        # Add gap row between sheets
        row = row_offset + max_r + 2

    for col_letter, w in col_widths.items():
        combined.column_dimensions[col_letter].width = w

    max_row_combined = combined.max_row
    max_col_combined = combined.max_column
    if max_row_combined >= 1 and max_col_combined >= 1:
        filter_range = f"A1:{get_column_letter(max_col_combined)}{max_row_combined}"
        combined.auto_filter.ref = filter_range
        print(f"Autofilter applied to range: {filter_range}")

    print(f"Saving: {output_file}")
    target_wb.save(output_file)
    print(f"Done! {len(sheets)} sheet(s) merged.")
    print(f"Output: {output_file}")


def main() -> int:
    input_file = sys.argv[1] if len(sys.argv) > 1 else "/a0/tmp/uploads/predloha.xlsx"
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    else:
        base, _ = os.path.splitext(input_file)
        output_file = f"{base}_combined_final.xlsx"

    if not os.path.exists(input_file):
        raise FileNotFoundError(input_file)

    merge_final(input_file=input_file, output_file=output_file)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

