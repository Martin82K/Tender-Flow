import importlib.util
import subprocess
import sys
import tempfile
import unittest
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Color, Font, NamedStyle, PatternFill, Protection, Side
from openpyxl.writer.theme import theme_xml


SCRIPT = Path(__file__).resolve().parents[2] / "server_py/excel_merge_tool/merge_final.py"
spec = importlib.util.spec_from_file_location("merge_final", SCRIPT)
merge_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(merge_module)


class ExcelMergeTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.input = Path(self.temp.name) / "Položky stavby.xlsx"
        self.output = Path(self.temp.name) / "výsledek.xlsx"
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Polozky"

    def merge(self):
        self.workbook.save(self.input)
        merge_module.merge_final(str(self.input), str(self.output))
        result = openpyxl.load_workbook(self.output)
        self.addCleanup(result.close)
        return result

    def assert_style_equal(self, source, target):
        for attribute in ("font", "fill", "border", "alignment", "number_format", "protection", "quotePrefix", "pivotButton"):
            with self.subTest(cell=source.coordinate, attribute=attribute):
                self.assertEqual(copy(getattr(source, attribute)), copy(getattr(target, attribute)))

    def test_preserves_rgb_font_and_fill_in_saved_workbook(self):
        cell = self.sheet["A1"]
        cell.value = "Polozka"
        cell.font = Font(name="Arial", bold=True, size=14, color="FF003366")
        cell.fill = PatternFill("solid", fgColor="FFCCDDFF")

        result = self.merge()

        self.assertEqual(result.active["B3"].value, "Polozka")
        self.assert_style_equal(cell, result.active["B3"])
        self.assertEqual(result.active["A1"].fill.fgColor.rgb, "00366092")
        self.assertEqual(result.active["A2"].font.color.rgb, "00FFFFFF")

    def test_registers_many_distinct_styles_in_target_workbook(self):
        for row in range(1, 16):
            cell = self.sheet.cell(row=row, column=1, value=row)
            cell.font = Font(name="Arial", size=10 + row, color=f"FF{row:02X}2244")
            cell.fill = PatternFill("solid", fgColor=f"FF55{row:02X}77")
            cell.border = Border(bottom=Side(style="thin", color=f"FF88{row:02X}AA"))

        result = self.merge()

        for row in range(1, 16):
            self.assert_style_equal(self.sheet.cell(row, 1), result.active.cell(row + 2, 2))

    def test_preserves_named_and_empty_cell_styles_without_changing_values(self):
        style = NamedStyle(
            name="Cena", font=Font(italic=True, color="FF116622"),
            fill=PatternFill("solid", fgColor="FFFFFF99"),
            border=Border(left=Side(style="double", color="FF123456")),
            alignment=Alignment(horizontal="right", wrap_text=True),
            number_format='#,##0.00 "CZK"', protection=Protection(locked=False, hidden=True),
        )
        self.sheet["A1"] = 1234.5
        self.sheet["A1"].style = style
        self.sheet["A1"].quotePrefix = True
        self.sheet["B1"].style = style
        self.sheet["C1"] = "Bez stylu"

        result = self.merge()

        self.assertEqual(result.active["B3"].value, 1234.5)
        self.assertIsNone(result.active["C3"].value)
        self.assert_style_equal(self.sheet["A1"], result.active["B3"])
        self.assert_style_equal(self.sheet["B1"], result.active["C3"])
        self.assertEqual(result.active["D3"].value, "Bez stylu")
        self.assertFalse(result.active["D3"].has_style)

    def test_preserves_workbook_theme_and_indexed_palette_for_cell_colors(self):
        self.workbook.loaded_theme = theme_xml.replace("4F81BD", "123456").encode("utf-8")
        palette = list(self.workbook._colors)
        palette[10] = "FF654321"
        self.workbook._colors = tuple(palette)
        self.sheet["A1"] = "Motiv"
        self.sheet["A1"].font = Font(color=Color(theme=4, tint=0.25))
        self.sheet["B1"] = "Paleta"
        self.sheet["B1"].fill = PatternFill("solid", fgColor=Color(indexed=10))

        result = self.merge()

        self.assertEqual(result.loaded_theme, self.workbook.loaded_theme)
        self.assertEqual(tuple(result._colors), tuple(self.workbook._colors))
        self.assert_style_equal(self.sheet["A1"], result.active["B3"])
        self.assert_style_equal(self.sheet["B1"], result.active["C3"])

    def test_preserves_merge_layout_formulas_dimensions_and_sheet_selection(self):
        self.sheet["A1"] = "Nadpis"
        self.sheet.merge_cells("A1:B1")
        self.sheet["A1"].font = Font(bold=True, color="FF112233")
        self.sheet["A2"] = 2
        self.sheet["B2"] = "=SUM(A2,3)"
        self.sheet.row_dimensions[1].height = 28
        self.sheet.column_dimensions["A"].width = 35
        self.workbook.create_sheet("Druhy")["A1"] = "Druha polozka"
        for name, state in (("Rekapitulace stavby", "visible"), ("Pokyny pro vyplnění", "visible"), ("Skryty", "hidden"), ("Velmi skryty", "veryHidden")):
            skip = self.workbook.create_sheet(name)
            skip["A1"] = "NEKOPIROVAT"
            skip.sheet_state = state

        result = self.merge()
        sheet = result.active

        self.assertEqual(result.sheetnames, ["Kombinovane"])
        self.assertEqual(sheet["A1"].value, "List")
        self.assertEqual(sheet["A2"].value, "=== Polozky ===")
        self.assertEqual(sheet["A2"].data_type, "s")
        self.assertEqual(sheet["A3"].value, "Polozky")
        self.assertEqual(sheet["B3"].value, "Nadpis")
        self.assertEqual(sheet["C4"].value, "=SUM(A2,3)")
        self.assertIn("B3:C3", {str(r) for r in sheet.merged_cells.ranges})
        self.assertEqual(sheet.row_dimensions[3].height, 28)
        self.assertEqual(sheet.column_dimensions["B"].width, 35)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:K7")
        values = [cell.value for row in sheet for cell in row]
        self.assertIn("Druha polozka", values)
        self.assertNotIn("NEKOPIROVAT", values)
        self.assert_style_equal(self.sheet["A1"], sheet["B3"])

    def test_cli_accepts_default_and_explicit_output_paths(self):
        self.sheet["A1"] = "CLI"
        self.sheet["A1"].font = Font(color="FF336699")
        self.workbook.save(self.input)
        default_output = self.input.with_name("Položky stavby_combined_final.xlsx")

        for arguments, output in (([str(self.input)], default_output), ([str(self.input), str(self.output)], self.output)):
            with self.subTest(output=output.name):
                process = subprocess.run([sys.executable, str(SCRIPT), *arguments], capture_output=True, text=True, timeout=30)
                self.assertEqual(process.returncode, 0, process.stderr)
                self.assertEqual(process.stderr, "")
                result = openpyxl.load_workbook(output)
                self.addCleanup(result.close)
                self.assertEqual(result.active["B3"].value, "CLI")
                self.assert_style_equal(self.sheet["A1"], result.active["B3"])

    def test_preserves_unstyled_data(self):
        self.sheet.append(["Polozka", 42, "=SUM(B1,1)"])

        result = self.merge()

        self.assertEqual([result.active.cell(3, col).value for col in range(2, 5)], ["Polozka", 42, "=SUM(B1,1)"])
        self.assertFalse(result.active["B3"].has_style)

    def test_cli_rejects_missing_input_without_creating_output(self):
        process = subprocess.run([sys.executable, str(SCRIPT), str(self.input), str(self.output)], capture_output=True, text=True, timeout=30)

        self.assertNotEqual(process.returncode, 0)
        self.assertIn("FileNotFoundError", process.stderr)
        self.assertFalse(self.output.exists())


if __name__ == "__main__":
    unittest.main()
